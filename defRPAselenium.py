from RPA.Browser.Selenium import Selenium;
from RPA.Excel.Application import Application
from RPA.Windows import Windows
from RPA.HTTP import HTTP
from RPA.Excel.Files import Files;
import time
import random
import os
from datetime import timedelta
import shutil
from datetime import date
from datetime import datetime
from pywinauto.keyboard import SendKeys
import logging
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s | %(name)s | %(levelname)s | %(message)s',
                    filename= 'log procesos' )

listSCRAPIADO= ([{
               'CUOTA':"",
               'NRO FOLIO':"", 
               'VALOR':"",
               'VENCIMIENTO':"",
                'TOTA A PAGAR':"",
                 }])

listSFormato= ([{
               'pathubicacion':"",
               'Nombre Solicitante':"", 
               'fecha':"",
               'gerente':"",
                'Rut':"",
                'Monto':"",
                'RUTtesoria':"",
                'Direccio':"",
                'Glosagasto':"",
                'Detallegasto':"",
                'CentroGestion':"",
                'Contribuciones':"",
                 }])

browser = Selenium()
library = Windows() 
lib = Files()
app = Application()
año="2023"
TxExcel=int(10)
RutExcel="-"
InmobiliariaExcel="-"

def valoresReporte(rut,inmobiliaria):
     global RutExcel
     global InmobiliariaExcel
     global TxExcel
     TxExcel+=1
     RutExcel=rut
     InmobiliariaExcel=inmobiliaria

     

def reportarError(mensaje,nomHoja,roles):
    # Cargar el archivo Excel
    fecha_actual=datetime.now()
    fecha_formateada = fecha_actual.strftime('%d/%m/%Y %H:%M:%S')
    archivo_excel = "Correo/Formato Reporte.xlsx"
    libro = openpyxl.load_workbook(archivo_excel)

    # Seleccionar la hoja en la que deseas escribir
    if nomHoja==True:
        hoja = libro.get_sheet_by_name("Errores")
    else:
        hoja = libro.get_sheet_by_name("Correctos")

    # Encontrar la última fila ocupada en una columna específica (por ejemplo, columna A)
    ultima_fila = hoja.max_row + 1

    # Lista de valores a insertar en las celdas
    valores_a_insertar = [str(TxExcel),str(fecha_formateada),str(RutExcel),str(InmobiliariaExcel),str(roles),str(mensaje)]
    # Escribir en las celdas vacías después de la última fila ocupada en la columna A
    for i, valor in enumerate(valores_a_insertar):
        hoja.cell(row=ultima_fila, column=i + 1, value=valor)

    # Guardar los cambios en el archivo
    libro.save(archivo_excel)
    libro.close()


def Pyasset(asset):
    lib.open_workbook("PyAsset\Config.xlsx")      #ubicacion del libro
    lib.read_worksheet("Variables")       #nombre de la hoja
    config=lib.read_worksheet_as_table(name='Variables',header=True, start=1).data
    for x in config:
        if x[0]==asset:
            exitdato= str(x[1])
        
            return exitdato

def openweb(u):

    duracion=timedelta(seconds=59)
    browser.set_selenium_page_load_timeout(duracion)
    try:
        browser.open_available_browser(u,browser_selection='firefox')
        browser.maximize_browser_window() 
    except:
         browser.reload_page()
    validacion= browser.get_text("//DIV[@class='dentro_letra'][text()='Contribuciones']")
    if validacion == 'Contribuciones': print("ingresando a "+validacion) 
    state_tgc_Inicio=True
 
    time.sleep(random.uniform(5,7)) 

def clickweb(elemento):
    time.sleep(random.uniform(1,2))
    browser.click_element(elemento)
    time.sleep(random.uniform(1,2))

def typeinputText(elemento,texto):
    time.sleep(random.uniform(1,2))
    browser.input_text(elemento,texto)
    time.sleep(random.uniform(1,2))

def obtenertabla(elemento,columna,celdas):
    time.sleep(random.uniform(1,2))
    browser.get_table_cell(locator=elemento,column=columna,row=celdas)
    time.sleep(random.uniform(1,2))

def obtenerTexto(elemento):
    time.sleep(random.uniform(1,2))
    browser.get_text(elemento)
    time.sleep(random.uniform(1,2))

def tiempoespera():
    time.sleep(random.uniform(20,30))

def cerraNavegador():
    browser.close_browser()
    print("----------------------proceso terminado----------------------")

def destacar(elemento):
    browser.highlight_elements(elemento)
    time.sleep(random.uniform(3,7))

def LOGconsulta(Región,Comuna,RolMatriz,Rol):
    print('----------------------Consultado-----------------------------')
    print('region = '+str(Región))
    print('Comuna = '+str(Comuna))
    print('Rol Matriz = '+str(RolMatriz))
    print('Rol = '+str(Rol))

def extraertablita():

    
    print(browser.get_text("//DIV[@id='example_info']/self::DIV"))
    
    
    scraping=browser.get_text("//TABLE[@id='example']")
    #recorrerFilasDescargas()
    print(scraping)
    return scraping
def filtrarCuota():
    #Noviembre trae 4 / Septiembre trae 3 / Junio trae 2 / Abril trae 1
    #Metodo para saber en que mes se esta corriendo y descargar los PDF´s que coincidan
    fecha_actual=datetime.now()
    fecha_formateada = fecha_actual.strftime("%B")
    anio_actual = fecha_actual.strftime('%Y')

    if fecha_formateada=="November":
        return "4-"+anio_actual
    elif fecha_formateada=="September":
        return "3-"+anio_actual
    elif fecha_formateada=="June":
        return "2-"+anio_actual
    elif fecha_formateada=="April":
        return "1-"+anio_actual
    elif fecha_formateada=="December":
        return "9-"+anio_actual
    else:
        print("Opción no valida")

def recorrerFilasDescargas(carpeta,driver,rol,hoja):
    row=-1
    consecutivo=0
    tabledata=txtscraping(carpeta)
    """filtro = filtroCuota(carpeta)
    intFiltro=int(filCuota[0:1])"""
    filtroCuota=filtrarCuota()
    intFiltro=int(filtroCuota[0:1])
    intAnioFiltro=int(filtroCuota[2:])
    #intFiltro=11
    for celda in tabledata:
        save=False
        row=row+1    
        try:
            CUOTA = celda.get('CUOTA')
            VALOR=  celda.get('VALOR')
            print(str(CUOTA))
            try:
                strCuota=str(CUOTA)
                cutCuota=strCuota[0:1]
                intCuota=int(cutCuota)
                anioCuota=int(strCuota[2:])
            except:
                intCuota=99 

            if  intCuota<=intFiltro or anioCuota<intAnioFiltro:
                consecutivo=consecutivo+1
                print("El consecutivo es " + str(consecutivo))
                #clickweb("//TABLE[@id='example']//tr["+str(row)+"]//td[3]")
                descarga=driver.find_element(By.XPATH,"//table[@id='example']//tr["+str(row)+"]//td[3]")
                descarga.click()
                try: creacioncarpetas(carpeta)
                except:pass
                while save==False:
                    try:
                         savepdf(carpeta,str(consecutivo),CUOTA,str(rol))
                         save=True
                    except:
                         save=False                 
        except:pass
        finally:pass
    reportarError("Cuotas rescatadas con exito",False,rol) 
            
def recorriendoFormatoSolicitud(carpeta,hoja):
    row=0
    
    tabledata=txtscraping(carpeta)
    try:
        for celda in tabledata:
            row=row+1  
            consecutivo=str(row)  
            CUOTA = celda.get('CUOTA')            
            VALOR=  celda.get('VALOR')
            si=str(CUOTA).find("-")

            if si == -1:                
                print("-----------------------------------------------------------------------")
            else:
                print("consultado hoja : "+hoja)
                print("consultado Cuota : "+str(CUOTA))
                print("consultado Monto : "+str(VALOR))
                row=int(row-1  )
                
               
                 
                row=row+1      
    except:
        pass
    finally:
            row=0
        
            tabledata=txtscraping(carpeta)
   
            for celda in tabledata:
                row=row+1  
                consecutivo=str(row)  
                CUOTA = celda.get('CUOTA')            
                VALOR=  celda.get('VALOR')
                si=str(CUOTA).find("-")

                if si == -1:                
                    print("-----------------------------------------------------------------------")
                else:
   
                    print("consultado hoja : "+hoja)
                    print("consultado Cuota : "+str(CUOTA))
                    print("consultado Monto : "+str(VALOR))
                    row=int(row-1  )
                
                    
                    row=row+1      
        
def validacion():
    validacion= browser.get_text("//DIV[@class='dentro_letra'][text()='Contribuciones']")
    if validacion == 'Contribuciones': print("ingresando a "+validacion) 
    return validacion
       
def navegacion(region,comuna,rol1,rol2,ruta,hoja):
    current_working_directory = os.getcwd()
    capsolver_extension_path = current_working_directory + "\\capsolver\\"
    chrome_service = Service()
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(
            "--load-extension={0}".format(capsolver_extension_path))


    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
    driver.set_page_load_timeout(60)
    try:
        driver.get("https://www.tesoreria.cl/ContribucionesPorRolWEB/muestraBusqueda?tipoPago=PortalContribPresencial")  
    except:
        driver.quit()

    driver.maximize_window()
    r=region.upper()
    elemento=driver.find_element(By.XPATH,"//select[@id='region']")
    elemento.click()
    time.sleep(3)
    elemento=driver.find_element(By.XPATH,"//option[text()='"+r+"']")
    elemento.click()
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//select[@id='comunas']")
    elemento.click()
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//option[text()='"+comuna+"']")
    elemento.click()
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//input[@id='rol']")
    elemento.send_keys(str(rol1))
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//input[@id='subRol']")
    elemento.send_keys(str(rol2))
    time.sleep(3)
    xpath = '//input[@id="btnRecaptchaV3Envio" and @class="boton g-recaptcha"]'
    elemento=driver.find_element(By.ID,"btnRecaptchaV3Envio")
    driver.execute_script("arguments[0].scrollIntoView();", elemento)
    elemento.click()
    time.sleep(5)
    try:
        capsolver = (WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/form/div"
                                                                                              "/div[4]/div/div[2]"))))
        capsolver.click()
        time.sleep(10)
    except:
        print("No funciono el captcha")
        

    espera=True
    cantidad=0
    #tiempoespera()
    while espera==True:
        cantidad+=1
        if cantidad==7:
            espera=False
            raise ValueError("Error")
        try:
            valida=driver.find_element(By.XPATH,"//div[@id='example_filter']")
            time.sleep(1)
            print("Ya cargó")
            espera=False
        except:
            espera=True
            print("-----------Cargando-------------")
            time.sleep(5)


    try: # Validando si la tabla funciona
        valida=driver.find_element(By.XPATH,"//td[@class='celdaContenido2  sorting_1'][text()='No se encontraron Deudas']")
        #valida=obtenerTexto("//TD[@class='celdaContenido2  sorting_1'][text()='No se encontraron Deudas']/self::TD")
        reportarError("No Presentaba Cuotas",False,str(rol1)+"-"+str(rol2))
        print("No hay registros de este Rol")        
    except:
        try:# proceso de consulta
                    #tabla = None
                    #tabla =extraertablita()
                    #export(ruta,tabla)
                    generarLogScrapingRol(driver,ruta)
                    pdfrol=str(rol1)+"-"+str(rol2)
                    recorrerFilasDescargas(ruta,driver,str(pdfrol),"nomInmobiliaria")
                    #recorrerFilasDescargas(ruta,tabla,str(pdfrol),False,nomInmobiliaria)
        except:# proceso de consulta reintento #1
            driver.quit()
            raise ValueError("Error")
            #cerraNavegador()

    finally:
         driver.quit()
         return driver
         #cerraNavegador() 

def generarLogScrapingRol(driver,carpeta):
    # Encuentra la tabla por su selector CSS (ajusta el selector según tu caso)
    tabla = driver.find_element(By.XPATH,"//table[@id='example']")

    # Encuentra todas las filas de la tabla
    filas = tabla.find_elements(By.TAG_NAME,'tr')

    # Abre un archivo de texto en modo escritura
    with open("Log Scraping/"+carpeta+".txt", 'w', encoding='utf-8') as archivo:

    # Itera a través de las filas y guarda el texto de las celdas en el archivo
        for fila in filas:
            celdas = fila.find_elements(By.TAG_NAME,'td')
            fila_datos = [celda.text for celda in celdas]
            fila_texto = ' '.join(map(str, fila_datos))  # Separar celdas con espacios
            archivo.write(fila_texto + '\n')  # Escribe la fila en el archivo



def savepdf(carpeta,consecutivo,cuota,rol):
 #Cerraw()
 base=Pyasset(asset="base")
 txt=base+carpeta
 switch=False
 alter=False
 #salida="Cupon de pago "+str(consecutivo)
 salida="Cupon de pago " + str(rol) + " " + str(cuota) + " consecutivo " + str(consecutivo)
 if str(consecutivo)=="1":
     consecutivo="1"
 
 
 try:
        file = open(txt+"\\"+salida)
        print(file) # File handler
        file.close()
 except:
    
    """library.click("name:imprimirAr")
    time.sleep(4.5)
    library.send_keys(keys="{CTRL}S")   
    time.sleep(4)"""

    origen=txt+"\\"+salida+".pdf"
    destino=txt+"\\"+"Cupon de pago "+str(rol)+" "+str(cuota)+".pdf"
    intentos=0
    while switch==False:
        if os.path.exists(origen):
             try:os.remove(origen)
             except:pass
        if os.path.exists(destino):
            print("El Archivo ya existe")
            switch=True
        else:
            library.click("name:imprimirAr")
            time.sleep(4.5)
            library.send_keys(keys="{CTRL}P")
            time.sleep(1)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{Enter}")
            time.sleep(0.5)
            library.send_keys(keys="g")
            time.sleep(0.5)
            library.send_keys(keys="{Enter}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{Enter}")
            time.sleep(0.5)


            library.send_keys(keys=txt)
            time.sleep(5)
            library.send_keys(keys="{Enter}")
            time.sleep(2)
            #library.send_keys(keys="{Alt}N")
            time.sleep(2)
            library.send_keys(keys="{CTRL}A")
            time.sleep(2)
            library.send_keys(keys=str(salida))
            time.sleep(3)
            library.send_keys(keys="{Enter}")
            time.sleep(2)
            switch=cambionombre(origen, destino,str(rol),str(intentos))
        intentos+=1
    library.send_keys(keys="{Enter}")
    time.sleep(1)
    library.send_keys(keys="{Esc}")
    time.sleep(1)
    library.click("name:imprimirAr")
    time.sleep(1)
    library.send_keys(keys="{Ctrl}W")



def txtscraping(carpeta):
    f = open('Log Scraping/' + carpeta + ".txt", "r")

    # Initialize a list to store the scraped data
    listSCRAPIADO = []
    scrp=[]
    # Loop through each line in the file
    for x in f:
        # Check if the line doesn't start with a space
        if x.find(" ") != 0:
            # Append the line to the 'scrp' list
            scrp.append(x)

    # Loop through each line in 'scrp'
    for u in scrp:
        # Find the index of the first space in the line
        final = u.find(" ")
        largo = len(u)
        Sumatoria = 0

        # Extract and clean data from the line
        CUOTA = str(u)[0:final]
        Sumatoria = Sumatoria + len(CUOTA) + 1

        dato = (str(u)[(Sumatoria):(largo - final)]).find(" ")
        VALOR = (str(u)[(Sumatoria):(Sumatoria + dato)]).replace(",", " ")
        Sumatoria = Sumatoria + len(VALOR) + 1

        dato = (str(u)[(Sumatoria):(largo - final)]).find(" ")
        NRO_FOLIO = (str(u)[(Sumatoria):(Sumatoria + dato)]).replace(",", " ")
        Sumatoria = Sumatoria + len(NRO_FOLIO) + 1

        dato = (str(u)[(Sumatoria):(largo - final)]).find(" ")
        VENCIMIENTO = (str(u)[(Sumatoria):(Sumatoria + dato)]).replace(",", " ")
        Sumatoria = Sumatoria + len(VENCIMIENTO) + 1

        dato = (str(u)[(Sumatoria):(largo - final)]).find(" ")
        TOTAPAGAR = (str(u)[(Sumatoria):(Sumatoria + dato)]).replace(",", " ")
        Sumatoria = Sumatoria + len(TOTAPAGAR) + 1

        # Create a dictionary with the extracted values and append it to the 'listSCRAPIADO' list
        listSCRAPIADO.append({
            'CUOTA': CUOTA,
            'NRO FOLIO': NRO_FOLIO,
            'VALOR': VALOR,
            'VENCIMIENTO': VENCIMIENTO,
            'TOTA A PAGAR': TOTAPAGAR,
        })

    # Close the file
    f.close()
         
      
    return listSCRAPIADO
    
def export(Carpeta,tabla):
     
     datosscrap=str(tabla) 
     outmensaje=datosscrap
     outmensaje=outmensaje.replace("VALOR"," ")
     outmensaje=outmensaje.replace("CUOTA"," ")
     outmensaje=outmensaje.replace("VALOR CUOTA"," " )
     outmensaje=outmensaje.replace("NRO FOLIO"," " )
     outmensaje=outmensaje.replace("VENCIMIENTO"," " )
     outmensaje=outmensaje.replace("TOTAL A PAGAR"," " )
     outmensaje=outmensaje.replace("EMAIL"," " )
     outmensaje=outmensaje.replace("DESCARGAR"," " )
     outmensaje=outmensaje.replace("""CUOTA
VALOR CUOTA
NRO FOLIO
VENCIMIENTO
TOTAL A PAGAR
EMAIL
DESCARGAR"""," " )

     try:
        file = open("Log Scraping/"+Carpeta+".txt","a")
        file.write(outmensaje)
        print(file) # File handler
        file.close()
       
     except:
        print("Archivo no existe se genera uno nuevo  "+ "Log Scraping/"+Carpeta+".txt")
        nom="Log Scraping/"+Carpeta+".txt"     
        f = open(nom, "a")
        f.write(outmensaje)
        f.close() 
                      

def cambionombre(origen, destino, rol,nIntentos):

        archivo = origen
        nombre_nuevo = destino
        print("archivo → "+ archivo )
        print("Destino → "+ nombre_nuevo)
        try:
            os.rename(archivo, nombre_nuevo)
            if os.path.exists(nombre_nuevo):
                 print("PDF guardado con exito")
                 return True
            else:
                print("Ocurrio un error al guardar el PDF, reintentando (Else)")
                library.send_keys(keys="{Enter}")
                library.send_keys(keys="{Esc}")
                reportarError("Fallo al guardar el PDF - Reintento N° "+str(nIntentos),True,rol)
                return False
        except:
             print("Ocurrio un error al guardar el PDF, reintentando (Except)")
             library.send_keys(keys="{Enter}")
             time.sleep(1)
             library.send_keys(keys="{Esc}")
             reportarError("Fallo al guardar el PDF - Reintento N° "+str(nIntentos),True,rol)
             return False

def Resumen():
    lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
    lib.read_worksheet("Resumen")                                              #nombre de la hoja
    dtresumen=lib.read_worksheet_as_table(name='Resumen',header=True, start=1).data
    return dtresumen

def master():
   lib.open_workbook("Data\Master.xlsx")      #ubicacion del libro
   lib.read_worksheet("Listado")       #nombre de la hoja
   DtMaster=lib.read_worksheet_as_table(name='Listado',header=True, start=1).data

   return DtMaster

def diligenciarResumen(h,carpeta):
    dtcon=txtscraping(carpeta)
   
       #ahora = datetime.now()
       #consulta=str(ahora.year)
    consulta="2023"
    
     
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
                       
            VALOR=  txt.get('VALOR')
            if str(CUOTA)[2:]==consulta : 
                cu = CUOTA             
                v = VALOR
        
                lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
                lib.read_worksheet("Resumen")                                              #nombre de la hoja
                libroresumen=lib.read_worksheet_as_table(name='Resumen',header=True, start=1).data    

                cantidad=lib.find_empty_row()

                #Ingresamos los valores 
                for celda in range(cantidad):
                
                    Numero=lib.get_cell_value(2+celda,"A")
                    if Numero==h:
                            lib.set_cell_value(2+celda,"E",str(v))
                            lib.set_cell_value(2+celda,"f","pago contribucciones "+str(cu))
                            lib.save_workbook() 
                                
def formatosolicitusd(h,carpeta):

    dtcon=txtscraping(carpeta)
    total=totalMacro(h)

    fecha_actual = datetime.now()

    fecha_formateada = fecha_actual.strftime('%d/%m/%Y')

    #ahora = datetime.now()
    #consulta=str(ahora.year)
    consulta="2023"
    
    siguiente=0 
    for txt in dtcon:
            siguiente=1+siguiente
            CUOTA = txt.get('CUOTA')                       
            VALOR=  txt.get('VALOR')
            if str(CUOTA)[2:]==consulta : 
                cu = CUOTA             
                v = VALOR 
                origen='Data\\Formato Solicitud Pago.xlsx'         
                destino="Formato Solicitud\\"+carpeta +" " +" Cuota " + str(cu) + " Formato Solicitud Pago.xlsx"
                #shutil.copy(origen,destino )

                    
                    
                datac=Resumen()

                for x in datac:
                    if x[0]==h:
                        
                        lib.open_workbook(origen)      
                        lib.read_worksheet("Solicitud")                                              
                        libroresumen=lib.read_worksheet_as_table(name='Solicitud',header=True, start=1).data
                        
                        lib.set_cell_value(8,"D",str(x[8]))

                        lib.set_cell_value(6,"H",str(fecha_formateada))
                        lib.set_cell_value(10,"D","Enrique Carrasco")
                        lib.set_cell_value(12,"D",str(x[3]))
                        lib.set_cell_value(12,"H",str(x[2]))  
                        lib.set_cell_value(14,"C",VALOR, fmt="0.00")
                        lib.set_cell_value(20,"C","Teatinos 28, Santiago")
                        lib.set_cell_value(22,"D","pago contribucciones "+str(CUOTA))
                        lib.set_cell_value(24,"D","pago contribucciones "+str(CUOTA))
                        lib.set_cell_value(26,"D",str(x[12]))
                        lib.set_cell_value(28,"D",str(x[12]))
                        lib.set_cell_value(30,"D",str("Contribucciones"))
                        lib.save_workbook(destino)
                        lib.close_workbook()
                break
            
def diligenciarhojas(h,carpeta,REGION,COMUNA,ROLMATRIZ,RUT,INMOBILIARIA,rol1,rol2):
    dtcon=txtscraping(carpeta)
    R=0
    celda=0

    for txt in dtcon:
         celda=1+celda

    print("el total de celdas es → "+str(celda))
    
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
            print(CUOTA)           
            VALOR=  txt.get('VALOR')
        
            lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
            lib.read_worksheet(str(h))                                                  #nombre de la hoja
            libroresumen=lib.read_worksheet_as_table(name=str(h),header=True, start=1).data    
            
            R=1+R 
                       
            lib.set_cell_value(6+R,"B",RUT) 
            lib.set_cell_value(6+R,"C",INMOBILIARIA)
            lib.set_cell_value(6+R,"D",REGION)
            lib.set_cell_value(6+R,"E",COMUNA)
            lib.set_cell_value(6,"H","Monto")
            lib.set_cell_value(5+R,"H",VALOR,fmt="0.00")
            lib.set_cell_value(6+R,"D",REGION)
            lib.set_cell_value(6+R,"E",COMUNA)
            lib.set_cell_value(6+R,"F",ROLMATRIZ)                   
            lib.save_workbook()
            
    print("el total de R es → "+str(R))
    R=0       
    lib.clear_cell_range("B16:H77")        
    for txt in dtcon:
            CUOTA = txt.get('CUOTA')                        
            VALOR=  txt.get('VALOR')
            R=1+R
            VO=lib.get_cell_value(5+R,"H")
            if VO is None:
                print(VO)
                lib.set_cell_value(5+R,"G"," ")
                lib.set_cell_value(5+R,"F"," ")
                lib.set_cell_value(5+R,"E"," ")
                lib.set_cell_value(5+R,"D"," ")
                lib.set_cell_value(5+R,"C"," ")
                lib.set_cell_value(5+R,"B"," ")
                break
            else:
                lib.set_cell_value(5+R,"G",CUOTA,fmt="0")

        
       
            
    #lib.set_cell_value(7+(R+2),"G","Total") 
    #lib.set_cell_formula("H17","=SUMA(H7:H16)",True)
   
    lib.set_cell_value(6,"H","Monto") 
    lib.save_workbook("Salida\\Resumen_Contribuciones_Terreno_2023.xlsx")#"Salida\\Resumen_Contribuciones_Terreno_2023.xlsx"
    lib.close_workbook ()      

def bakup():
     
     print("Realizamos el bakup")
     origen='Data\\BACKUP\\Resumen_Contribuciones_Terreno_2023.xlsx'         
     destino="Data\\Resumen_Contribuciones_Terreno_2023.xlsx"
     shutil.copy(origen,destino )

def creacioncarpetas (carpeta):

    os.mkdir('PDF/'+carpeta)    
    print("creacion de carpetas  PDF/"+carpeta) 

def Macros (h):
    lib.open_workbook("Data\Macro TGR.xlsm")      
    lib.read_worksheet("MACRO")                                                                     
    libroresumen=lib.read_worksheet_as_table(name="MACRO",header=True, start=1).data 

    lib.set_cell_value(3,"B",str(h))
    lib.save_workbook()
    lib.close_workbook()
    

    app.open_application(visible=True)
    time.sleep(10)
    try:
         library.click("name:Cerrar")
    except:
         logging.info("No encontro licencia de excel vencida")
         pass
    
    logging.info("Preparando Macro ingresando")
    app.open_workbook('Data\Macro TGR.xlsm')
    app.set_active_worksheet(sheetname="MACRO")
    time.sleep(5)
    app.run_macro("Main")
    logging.info("Macro ejecutada con exito")
    time.sleep(5)
    app.save_excel()
    app.quit_application()

def totalMacro(h):
    lib.open_workbook("Data\Resumen_Contribuciones_Terreno_2023.xlsx")      
    lib.read_worksheet(h)                                                                     
    libroresumen=lib.read_worksheet_as_table(name=str(h),header=True, start=1).data 

    TOTAL =lib.get_cell_value(20,"H")

    lib.save_workbook()
    lib.close_workbook()
    return TOTAL

def formatoTotal(h,carpeta):


    totalv=int(totalMacro(h))
    print(str(totalv))
    
    dtcon=txtscraping(carpeta)

    fecha_actual = datetime.now()

    fecha_formateada = fecha_actual.strftime('%d/%m/%Y')
    

    #ahora = datetime.now()
    #consulta=str(ahora.year)
    consulta="2023"
    
      
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
                       
            VALOR=  txt.get('VALOR')
            if str(CUOTA)[2:]==consulta : 
                cu = CUOTA             
                v = VALOR 
       
                destino="Formato Solicitud\\"+carpeta +" " +" Cuota " + str(cu) + " Formato Solicitud Pago.xlsx"
                #shutil.copy(origen,destino )

                datac=Resumen()

                for x in datac:
                    if x[0]==h:

                        lib.open_workbook(destino)      
                        lib.read_worksheet("Solicitud")                                                                     
                        libroresumen=lib.read_worksheet_as_table(name='Solicitud',header=True, start=1).data 

                    
                        lib.set_cell_value(14,"C",int(totalv), fmt="0.00")
                        

            

    lib.save_workbook()
    lib.close_workbook()

def fGuardar(h,carpeta):

    dtcon=txtscraping(carpeta)

    fecha_actual = datetime.now()

    fecha_formateada = fecha_actual.strftime('%d/%m/%Y')

    #ahora = datetime.now()
    #consulta=str(ahora.year)
    consulta="2023"
    
      
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
                       
            VALOR=  txt.get('VALOR')
            if str(CUOTA)[2:]==consulta : 
                cu = CUOTA             
                v = VALOR 
     
    destino="Formato Solicitud\\"+carpeta +" " +" Cuota " + str(cu) + ".xlsm"

    lib.open_workbook("Data\Resumen_Contribuciones_Terreno_2023.xlsm")      
    lib.read_worksheet("Solicitud")                                                                     
    libroresumen=lib.read_worksheet_as_table(name='Solicitud',header=True, start=1).data 

    lib.set_cell_value(1,"k",int(h))

    lib.save_workbook(destino)
    lib.close_workbook()

def ResumenFinal ():
     lib.open_workbook('Data\Resumen_Contribuciones_Terreno_2023.xlsx')        #ubicacion del libro
     lib.read_worksheet('Resumen')       #nombre de la hoja
     lista=lib.read_worksheet_as_table(name='Resumen',header=True, start=1).data

     ultimaFila= lib.find_empty_row()

     for celda in range(ultimaFila):
         TOTAL= lib.get_cell_value(2+int(celda),"E")
         if TOTAL == "=+'1'!$H$9":
              print("True "+str(TOTAL))
         else:
              print("false "+str(TOTAL))
              HOJA=lib.get_cell_value(2+int(celda),"A")
              
              lib.read_worksheet(str(HOJA))
              tablaTotal=lib.get_cell_value(20,"H")
             

              lib.read_worksheet('Resumen')
              lib.set_cell_value(2+int(celda),"E",int(tablaTotal))



              lib.save_workbook()
              lib.close_workbook()

def limpiarResumen():
     
     
     lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
     lib.read_worksheet("94")                                                 #nombre de la hoja
     libroresumen=lib.read_worksheet_as_table(name="94",header=True, start=1).data  
      

     lib.clear_cell_range("G7:G1000")
     rango="B{}:H{}"
     
     #Comparaciones 
     item1=lib.get_cell_value(7,"H")
     item2=lib.get_cell_value(8,"H")
     item3=lib.get_cell_value(9,"H")
     item4=lib.get_cell_value(10,"H")
     item5=lib.get_cell_value(11,"H")
     item6=lib.get_cell_value(12,"H")
     item7=lib.get_cell_value(13,"H")
     item8=lib.get_cell_value(14,"H")
     item9=lib.get_cell_value(16,"H")
     item10=lib.get_cell_value(17,"H")
      
     busquedad=0
     for x in range(1000):           
            Cels=str(x+8)
                   
            if item1==lib.get_cell_value(7,"H"):
               busquedad=1+busquedad   
            elif busquedad>1:      
               lib.clear_cell_range(rango.format(Cels,Cels))

     busquedad=0
     for x in range(1000):           
            Cels=str(x+9)
                   
            if item1==lib.get_cell_value(8,"H"):
               busquedad=1+busquedad   
            elif busquedad>1:      
               lib.clear_cell_range(rango.format(Cels,Cels))




     lib.save_workbook()
     lib.close_workbook()

def salida():
     
     print("Realizamos la salida ")
     origen='Data\Resumen_Contribuciones_Terreno_2023.xlsx'         
     destino="Salida\Resumen_Contribuciones_Terreno_2023.xlsx"
     shutil.copy(origen,destino )

def cerrarinicio():
     try:
          browser.close_browser()
     except:
      pass

def pdfsaturado(carpeta,consecutivo,cuota,rol):
    try:
        
        
        library.click("name:No")
        time.sleep(2)
        library.send_keys(keys="{Alt}N")
        time.sleep(2)
        library.send_keys(keys=carpeta+" "+cuota+" "+rol+consecutivo+1)
        time.sleep(3)
        library.send_keys(keys="{Enter}")
        time.sleep(1)
        library.send_keys(keys="{Enter}")
        consecutivo=consecutivo+1
        browser.close_browser()
        browser.close_all_browsers()
        browser.close_window()
        
    except:
         print("pdf no esta saturado")
         browser.close_browser()
         browser.close_all_browsers()
         browser.close_window()
         pass

def filtroCuota(carpeta):
    
    f=open('Log Scraping/'+carpeta+'.txt',"r")
    fecha_actual = datetime.now()
    fecha_formateada = fecha_actual.strftime('%Y')
    Cuo=[]
    añocuo=[]
    a=0
    subtotal=0
    for a in f:
            if a.__contains__(fecha_formateada):
                Cuo.append(a[0:2].replace("-",""))
                añocuo.append(a[2:7].replace("-",""))
    mincuo=max(Cuo)  
    añomax=max(añocuo) 
    print("Filtra cuota a pagar")  
    print(mincuo)
    print(añomax)
    filtrocuota=mincuo+"-"+añomax
    f=str(filtrocuota).strip()
    print(f)
    return f


def viejorecorrerFilasDescargas(carpeta,scraping,rol,hoja):
   
    row=0
    tabledata=txtscraping(carpeta)
    filtro = filtroCuota(carpeta)

    for celda in tabledata:
        
        row=row+1  
        consecutivo=str(row)     
        try:
                CUOTA = celda.get('CUOTA')
                VALOR=  celda.get('VALOR')
                
                si=str(CUOTA).find("-")
    
                if si == -1:
                    print("la cuota no es visible ")
                else:
                    row=int(row-1)
                    consecutivo=str(row)
                    obtenerTexto("//TABLE[@id='example']//tr["+str(row)+"]//td[3]")
                    FOLIO=obtenerTexto("//TABLE[@id='example']//tr["+str(row)+"]//td[3]")
                    print("El consecutivo es " + str(consecutivo ))
                    clickweb("//TABLE[@id='example']//tr["+str(row)+"]//td[3]")

                    try:
                        creacioncarpetas (carpeta)
                    except:
                         pass
                    
                    
                    savepdf(carpeta,str(consecutivo ),CUOTA,str(rol))
                    row=row+1 
                   
        except:
             pass
        finally:
            pass


def Cerraw():
     try:library.click("name:Cerrar")
     except:library.click("name:Cancelar")
     finally:pass


